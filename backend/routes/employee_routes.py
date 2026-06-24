import os
import csv
from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, session, current_app
from werkzeug.utils import secure_filename
from openpyxl import load_workbook
from backend.database import db
from backend.models import Employee, User, Role, Department, LeaveRequest, Attendance, Payroll, Document, PerformanceReview
from backend.auth import login_required, role_required, token_required, api_role_required
from backend.services.audit_service import log_action
from backend.services.email_service import send_welcome_email
from datetime import datetime

employee_bp = Blueprint('employee', __name__)

def generate_employee_id():
    """Generates next sequential employee code: EMP1001, EMP1002..."""
    count = db.session.query(Employee).count()
    return f"EMP{1001 + count}"

# ----------------- Web Views -----------------

@employee_bp.route('/employees')
@login_required
@role_required('Super Admin', 'HR Manager', 'Team Lead')
def list_view():
    departments = Department.query.all()
    # List managers for dropdown
    managers = Employee.query.filter(Employee.designation.ilike('%manager%') | Employee.designation.ilike('%lead%')).all()
    return render_template('employees.html', departments=departments, managers=managers)

@employee_bp.route('/employees/<int:emp_id>')
@login_required
def detail_view(emp_id):
    # Enforce access control: Employees can only view their own profile, others require permissions
    if session['role_name'] == 'Employee':
        curr_emp = Employee.query.filter_by(user_id=session['user_id']).first()
        if not curr_emp or curr_emp.id != emp_id:
            flash("Access Denied: You can only view your own profile.", "danger")
            return redirect(url_for('auth.dashboard_view'))

    emp = Employee.query.get_or_404(emp_id)
    departments = Department.query.all()
    managers = Employee.query.filter(Employee.id != emp_id).all()
    
    # Details for child tabs
    attendance = Attendance.query.filter_by(employee_id=emp_id).order_by(Attendance.date.desc()).all()
    leaves = LeaveRequest.query.filter_by(employee_id=emp_id).order_by(LeaveRequest.start_date.desc()).all()
    payroll = Payroll.query.filter_by(employee_id=emp_id).order_by(Payroll.year.desc(), Payroll.month.desc()).all()
    documents = Document.query.filter_by(employee_id=emp_id).order_by(Document.uploaded_at.desc()).all()
    reviews = PerformanceReview.query.filter_by(employee_id=emp_id).order_by(PerformanceReview.created_at.desc()).all()
    
    return render_template(
        'employee_detail.html',
        employee=emp,
        departments=departments,
        managers=managers,
        attendance=attendance,
        leaves=leaves,
        payroll=payroll,
        documents=documents,
        reviews=reviews
    )

# ----------------- REST APIs -----------------

@employee_bp.route('/api/employees', methods=['GET'])
@token_required
def api_list_employees():
    """Paginated list of employees with search and filters"""
    query = Employee.query
    
    # Filters
    search = request.args.get('search')
    dept_id = request.args.get('department_id')
    status = request.args.get('status')
    gender = request.args.get('gender')
    emp_type = request.args.get('employment_type')
    
    if search:
        search_filter = f"%{search}%"
        query = query.filter(
            db.or_(
                Employee.first_name.ilike(search_filter),
                Employee.last_name.ilike(search_filter),
                Employee.email.ilike(search_filter),
                Employee.employee_id.ilike(search_filter),
                Employee.designation.ilike(search_filter)
            )
        )
    if dept_id:
        query = query.filter(Employee.department_id == int(dept_id))
    if status:
        active_status = status.lower() == 'active'
        query = query.filter(Employee.active_status == active_status)
    if gender:
        query = query.filter(Employee.gender.ilike(gender))
    if emp_type:
        query = query.filter(Employee.employment_type == emp_type)
        
    # Pagination
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 10, type=int)
    paginated = query.paginate(page=page, per_page=per_page, error_out=False)
    
    return jsonify({
        'employees': [emp.to_dict() for emp in paginated.items],
        'total': paginated.total,
        'pages': paginated.pages,
        'current_page': paginated.page
    }), 200

@employee_bp.route('/api/employees/<int:emp_id>', methods=['GET'])
@token_required
def api_get_employee(emp_id):
    emp = Employee.query.get_or_404(emp_id)
    # Check permissions
    if request.user_role == 'Employee' and request.employee_id != emp.id:
        return jsonify({'message': 'Access forbidden'}), 403
    return jsonify(emp.to_dict()), 200

@employee_bp.route('/api/employees', methods=['POST'])
@token_required
@api_role_required('Super Admin', 'HR Manager')
def api_create_employee():
    """Create a new employee record and associated credentials user account"""
    # Parse normal fields or form data (for profile image support)
    data = request.form.to_dict() if request.form else request.get_json() or {}
    
    email = data.get('email')
    if not email:
        return jsonify({'message': 'Email is required'}), 400
        
    if User.query.filter_by(email=email).first() or Employee.query.filter_by(email=email).first():
        return jsonify({'message': 'User with this email already exists'}), 400

    # Handle image upload
    profile_photo_path = 'uploads/profiles/default.png'
    if 'profile_photo' in request.files:
        file = request.files['profile_photo']
        if file and file.filename != '':
            filename = secure_filename(f"{uuid_name()}_{file.filename}")
            upload_dir = os.path.join(current_app.root_path, 'static', 'uploads', 'profiles')
            os.makedirs(upload_dir, exist_ok=True)
            file.save(os.path.join(upload_dir, filename))
            profile_photo_path = f"uploads/profiles/{filename}"

    # Determine default user Role (matches designation if Lead/Manager)
    designation = data.get('designation', '').lower()
    role_name = 'Employee'
    if 'manager' in designation:
        role_name = 'HR Manager'
    elif 'lead' in designation or 'supervisor' in designation:
        role_name = 'Team Lead'
        
    role = Role.query.filter_by(name=role_name).first()

    # Create associated login User
    temp_pass = "Welcome@123"
    user = User(email=email, role_id=role.id, is_active=True, email_verified=False)
    user.set_password(temp_pass)
    db.session.add(user)
    db.session.flush() # Grab user.id

    # Create Employee
    emp_code = generate_employee_id()
    dob = datetime.strptime(data.get('date_of_birth'), '%Y-%m-%d').date() if data.get('date_of_birth') else None
    joining = datetime.strptime(data.get('joining_date'), '%Y-%m-%d').date() if data.get('joining_date') else date.today()

    new_emp = Employee(
        user_id=user.id,
        employee_id=emp_code,
        first_name=data.get('first_name'),
        last_name=data.get('last_name'),
        email=email,
        mobile_number=data.get('mobile_number'),
        emergency_contact=data.get('emergency_contact'),
        gender=data.get('gender'),
        date_of_birth=dob,
        blood_group=data.get('blood_group'),
        marital_status=data.get('marital_status'),
        address=data.get('address'),
        city=data.get('city'),
        state=data.get('state'),
        country=data.get('country'),
        postal_code=data.get('postal_code'),
        department_id=int(data.get('department_id')) if data.get('department_id') else None,
        designation=data.get('designation'),
        salary=float(data.get('salary', 0.0)),
        joining_date=joining,
        manager_id=int(data.get('manager_id')) if data.get('manager_id') else None,
        profile_photo=profile_photo_path,
        aadhaar_number=data.get('aadhaar_number'),
        pan_number=data.get('pan_number'),
        employment_type=data.get('employment_type', 'Full-Time'),
        active_status=True
    )
    
    db.session.add(new_emp)
    db.session.commit()
    
    # Send login credentials to employee
    send_welcome_email(f"{new_emp.first_name} {new_emp.last_name}", email, temp_pass)
    log_action(f"Created Employee record {emp_code}", request.user_id)
    
    return jsonify(new_emp.to_dict()), 201

@employee_bp.route('/api/employees/<int:emp_id>', methods=['PUT'])
@token_required
@api_role_required('Super Admin', 'HR Manager')
def api_update_employee(emp_id):
    emp = Employee.query.get_or_404(emp_id)
    data = request.form.to_dict() if request.form else request.get_json() or {}
    
    # Update fields
    if 'first_name' in data: emp.first_name = data['first_name']
    if 'last_name' in data: emp.last_name = data['last_name']
    if 'mobile_number' in data: emp.mobile_number = data['mobile_number']
    if 'emergency_contact' in data: emp.emergency_contact = data['emergency_contact']
    if 'gender' in data: emp.gender = data['gender']
    if 'blood_group' in data: emp.blood_group = data['blood_group']
    if 'marital_status' in data: emp.marital_status = data['marital_status']
    if 'address' in data: emp.address = data['address']
    if 'city' in data: emp.city = data['city']
    if 'state' in data: emp.state = data['state']
    if 'country' in data: emp.country = data['country']
    if 'postal_code' in data: emp.postal_code = data['postal_code']
    if 'designation' in data: emp.designation = data['designation']
    if 'employment_type' in data: emp.employment_type = data['employment_type']
    if 'aadhaar_number' in data: emp.aadhaar_number = data['aadhaar_number']
    if 'pan_number' in data: emp.pan_number = data['pan_number']
    
    if 'department_id' in data and data['department_id']: 
        emp.department_id = int(data['department_id'])
    if 'manager_id' in data:
        emp.manager_id = int(data['manager_id']) if data['manager_id'] else None
    if 'salary' in data: 
        emp.salary = float(data['salary'])
    if 'active_status' in data:
        status_val = str(data['active_status']).lower() in ['true', '1', 'on']
        emp.active_status = status_val
        if emp.user:
            emp.user.is_active = status_val
            
    if 'date_of_birth' in data and data['date_of_birth']:
        emp.date_of_birth = datetime.strptime(data['date_of_birth'], '%Y-%m-%d').date()
    if 'joining_date' in data and data['joining_date']:
        emp.joining_date = datetime.strptime(data['joining_date'], '%Y-%m-%d').date()

    # Image upload
    if 'profile_photo' in request.files:
        file = request.files['profile_photo']
        if file and file.filename != '':
            filename = secure_filename(f"{uuid_name()}_{file.filename}")
            upload_dir = os.path.join(current_app.root_path, 'static', 'uploads', 'profiles')
            os.makedirs(upload_dir, exist_ok=True)
            file.save(os.path.join(upload_dir, filename))
            emp.profile_photo = f"uploads/profiles/{filename}"

    db.session.commit()
    log_action(f"Updated Employee profile {emp.employee_id}", request.user_id)
    return jsonify(emp.to_dict()), 200

@employee_bp.route('/api/employees/<int:emp_id>', methods=['DELETE'])
@token_required
@api_role_required('Super Admin')
def api_delete_employee(emp_id):
    emp = Employee.query.get_or_404(emp_id)
    code = emp.employee_id
    
    # Associated user is cascade deleted due to foreign key constraints,
    # but let's delete explicitly if needed.
    if emp.user:
        db.session.delete(emp.user)
    else:
        db.session.delete(emp)
        
    db.session.commit()
    log_action(f"Deleted Employee {code}", request.user_id)
    return jsonify({'message': 'Employee deleted successfully'}), 200

@employee_bp.route('/api/employees/bulk-delete', methods=['POST'])
@token_required
@api_role_required('Super Admin')
def api_bulk_delete():
    data = request.get_json() or {}
    ids = data.get('ids', [])
    if not ids:
        return jsonify({'message': 'No IDs provided'}), 400
        
    employees = Employee.query.filter(Employee.id.in_(ids)).all()
    count = 0
    for emp in employees:
        if emp.user:
            db.session.delete(emp.user)
        else:
            db.session.delete(emp)
        count += 1
        
    db.session.commit()
    log_action(f"Bulk deleted {count} employees", request.user_id)
    return jsonify({'message': f'Successfully deleted {count} employees'}), 200

@employee_bp.route('/api/employees/import', methods=['POST'])
@token_required
@api_role_required('Super Admin', 'HR Manager')
def api_bulk_import():
    """Bulk import employees from CSV or XLSX sheets"""
    if 'file' not in request.files:
        return jsonify({'message': 'No file uploaded'}), 400
        
    file = request.files['file']
    if not file or file.filename == '':
        return jsonify({'message': 'No file selected'}), 400
        
    ext = file.filename.rsplit('.', 1)[-1].lower()
    imported_count = 0
    
    try:
        temp_path = os.path.join(current_app.root_path, 'static', 'uploads', f"temp_import_{uuid_name()}.{ext}")
        os.makedirs(os.path.dirname(temp_path), exist_ok=True)
        file.save(temp_path)

        rows = []
        if ext == 'csv':
            with open(temp_path, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                rows = list(reader)
        elif ext in ['xlsx', 'xls']:
            wb = load_workbook(temp_path)
            sheet = wb.active
            headers = [cell.value for cell in sheet[1]]
            for r in range(2, sheet.max_row + 1):
                row_val = [sheet.cell(row=r, column=c).value for c in range(1, len(headers) + 1)]
                if any(row_val):
                    rows.append(dict(zip(headers, row_val)))
        
        # Process rows
        role_emp = Role.query.filter_by(name='Employee').first()
        
        for r in rows:
            email = r.get('email')
            if not email or User.query.filter_by(email=email).first() or Employee.query.filter_by(email=email).first():
                continue # Skip existing or empty emails
                
            # Create user login account
            temp_pass = "Welcome@123"
            user = User(email=email, role_id=role_emp.id, is_active=True)
            user.set_password(temp_pass)
            db.session.add(user)
            db.session.flush()

            # Parse optional dates
            dob = datetime.strptime(str(r.get('date_of_birth')), '%Y-%m-%d').date() if r.get('date_of_birth') else None
            joining = datetime.strptime(str(r.get('joining_date')), '%Y-%m-%d').date() if r.get('joining_date') else date.today()

            # Find department or map to first
            dept_name = r.get('department')
            dept_id = None
            if dept_name:
                dept = Department.query.filter(Department.department_name.ilike(dept_name)).first()
                if dept:
                    dept_id = dept.id

            emp = Employee(
                user_id=user.id,
                employee_id=generate_employee_id(),
                first_name=r.get('first_name', 'Imported'),
                last_name=r.get('last_name', 'Employee'),
                email=email,
                mobile_number=r.get('mobile_number'),
                gender=r.get('gender'),
                date_of_birth=dob,
                joining_date=joining,
                department_id=dept_id or 1,
                designation=r.get('designation', 'Associate'),
                salary=float(r.get('salary', 0.0) or 0.0),
                profile_photo='uploads/profiles/default.png',
                employment_type=r.get('employment_type', 'Full-Time'),
                active_status=True
            )
            db.session.add(emp)
            db.session.flush()
            
            # Send welcome email
            send_welcome_email(f"{emp.first_name} {emp.last_name}", email, temp_pass)
            imported_count += 1

        db.session.commit()
        if os.path.exists(temp_path):
            os.remove(temp_path)
            
        log_action(f"Imported {imported_count} employees via sheet upload", request.user_id)
        return jsonify({'message': f'Successfully imported {imported_count} employees'}), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'message': f'Error parsing sheet: {str(e)}'}), 500

def uuid_name():
    import uuid
    return str(uuid.uuid4())[:8]
